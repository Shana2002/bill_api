from flask import request
from app.init import app
from app.models.bill_data_model import BillData 
from app.models.template_model import Template
from openpyxl import Workbook,load_workbook
from app.models.convert_pdf_model import Convert
from app.controllers.send_message import send_whatsapp
from app.controllers.send_email import SendMail
import shutil
import datetime
import os
from dotenv import load_dotenv

load_dotenv
baseurl= os.getenv("BASE_URL")
@app.post('/generate_bill/')
def generate_bill():
    data = request.get_json()
    bill = BillData(data)
    template = Template(int(bill.template))

    file_name  = f"{bill.invoice_num}_{bill.customer_name}"

    shutil.copyfile(f"assets/templates_excel/template{int(bill.template)}.xlsx", f"assets/output_excel/{file_name}.xlsx")
    workbook = load_workbook(f"assets/output_excel/{file_name}.xlsx")
    sheet = workbook.active

    # Assign company details
    sheet[template.company_name] = bill.company_name
    sheet[template.email] = f'emai: {bill.company_email}'
    sheet[template.contact] = f'Contact: {bill.company_contact}'

    # Assign Customer Details
    sheet[template.customer] = f'Customer: {bill.customer_name}'
    sheet[template.cus_contact] = f'Customer conact: {bill.customer_contact}'

    # Assign bill varibales
    total = 0
    i =0

    # Assign items
    items = bill.items
    start_row = template.start

    # check items length
    if len(items) != 0:
        for item in items:
            if template.no_col_show:
                sheet.cell(row=start_row,column=template.no_col).value = i
            
            sheet.cell(row=start_row,column=template.item_col).value = item["item"]
            sheet.cell(row=start_row,column=template.price_col).value = item["price"]
            sheet.cell(row=start_row,column=template.qty_col).value = item["qty"]
            sheet.cell(row=start_row,column=template.price_col).value = item["price"]*item["qty"]
            total += float(float(item["price"])*float(item["qty"]))
            start_row += 1


    sheet[template.total] = total
    sheet[template.discount]= bill.discount
    sheet[template.sub_totla]= total-(total*float(bill.discount)/100)
    # save excel
    workbook.save(f"assets/output_excel/{file_name}.xlsx")
    
    new_pdf = Convert(f"{file_name}")
    new_pdf.convert_pdf()


    print(send_whatsapp(file_name)) 
    mail = SendMail(bill,file_name)
    mail.send_email()

    return ({
        "pdf":f"{baseurl}/get/output_pdf/{file_name}.pdf",
        "excel":f"{baseurl}/get/output_excel/{file_name}.xlsx"
        })

